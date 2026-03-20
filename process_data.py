"""Process all EXPO Excel files into a single JSON data file for the dashboard."""
import json
import os
import glob
from openpyxl import load_workbook
from datetime import datetime

BASE = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.dirname(BASE)
OUT = os.path.join(BASE, "public", "expo-data.json")


def read_2023():
    path = glob.glob(os.path.join(DATA_DIR, "2023", "*.xlsx"))[0]
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(h).strip() if h else "" for h in rows[0]]
    records = []
    for row in rows[1:]:
        r = dict(zip(headers, row))
        # Parse date
        checkin = r.get("CheckIn Date", "")
        if isinstance(checkin, datetime):
            date_str = checkin.strftime("%Y-%m-%d")
        elif isinstance(checkin, str) and checkin:
            try:
                date_str = datetime.strptime(checkin.split(" ")[0], "%d/%m/%Y").strftime("%Y-%m-%d")
            except Exception:
                date_str = str(checkin)
        else:
            date_str = ""

        company = str(r.get("Company Name", "") or "").strip()
        reg_type = str(r.get("Registration Type", "") or "").strip()
        venue = str(r.get("Venue", "") or "").strip()
        main_name = str(r.get("Main Attendee", "") or "").strip()
        surname = str(r.get("Surname", "") or "").strip()

        # Count attendees (main + up to 3 additional)
        attendee_count = 1
        for col in ["Attendee1", "Attendee2", "Attendee3"]:
            val = r.get(col)
            if val and str(val).strip():
                attendee_count += 1

        records.append({
            "year": 2023,
            "date": date_str,
            "venue": venue,
            "company": company,
            "firstName": main_name,
            "lastName": surname,
            "registrationType": reg_type,
            "accountManager": "",
            "accountManagerState": "",
            "dealerCode": "",
            "attendeeCount": attendee_count,
        })
    wb.close()
    return records


def read_2024_2025(year):
    folder = os.path.join(DATA_DIR, str(year))
    files = glob.glob(os.path.join(folder, "*.xlsx"))
    records = []
    for path in files:
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        headers = [str(h).strip() if h else "" for h in rows[0]]
        for row in rows[1:]:
            r = dict(zip(headers, row))
            entry_date = r.get("Entry Date", "")
            if isinstance(entry_date, datetime):
                date_str = entry_date.strftime("%Y-%m-%d")
            elif isinstance(entry_date, str) and entry_date:
                date_str = str(entry_date)
            else:
                date_str = ""

            venue_raw = str(r.get("Select Venue", "") or "").strip()
            # Extract city name from venue string
            venue = venue_raw
            for city in ["Brisbane", "Sydney", "Melbourne", "Perth", "Adelaide"]:
                if city.lower() in venue_raw.lower():
                    venue = city
                    break

            company = str(r.get("Company Name", "") or "").strip()
            reg_type = str(r.get("Registration Type", "") or "").strip()
            am = str(r.get("Account Manager", "") or "").strip()
            am_state = str(r.get("Account Manager State", "") or "").strip()
            dealer_code = str(r.get("Dealer Code", "") or "").strip()
            first_name = str(r.get("First Name", "") or "").strip()
            last_name = str(r.get("Last Name", "") or "").strip()

            records.append({
                "year": year,
                "date": date_str,
                "venue": venue,
                "company": company,
                "firstName": first_name,
                "lastName": last_name,
                "registrationType": reg_type,
                "accountManager": am,
                "accountManagerState": am_state,
                "dealerCode": dealer_code,
                "attendeeCount": 1,
            })
        wb.close()
    return records


def compute_analytics(all_records):
    """Compute all analytics from raw records."""

    # --- Basic counts by year and venue ---
    years = sorted(set(r["year"] for r in all_records))
    by_year = {}
    for y in years:
        recs = [r for r in all_records if r["year"] == y]
        by_year[y] = recs

    # --- Year-over-year summary ---
    companies_by_year = {}
    for y in years:
        companies_by_year[y] = set(r["company"].lower() for r in by_year[y] if r["company"])

    yoy = []
    all_prev_companies = set()
    for y in years:
        current = companies_by_year[y]
        returning = current & all_prev_companies if all_prev_companies else set()
        new = current - all_prev_companies if all_prev_companies else current
        total_attendees = sum(r["attendeeCount"] for r in by_year[y])
        yoy.append({
            "year": str(y),
            "totalAttendees": total_attendees,
            "uniqueCompanies": len(current),
            "returningCompanies": len(returning),
            "newCompanies": len(new),
        })
        all_prev_companies |= current

    # --- Venue breakdown per year ---
    venue_data = []
    for y in years:
        venues = {}
        for r in by_year[y]:
            v = r["venue"] or "Unknown"
            venues[v] = venues.get(v, 0) + r["attendeeCount"]
        for v, count in sorted(venues.items(), key=lambda x: -x[1]):
            venue_data.append({"year": str(y), "venue": v, "attendees": count})

    # --- Registration type breakdown ---
    reg_type_data = []
    for y in years:
        types = {}
        for r in by_year[y]:
            t = r["registrationType"] or "Unknown"
            types[t] = types.get(t, 0) + r["attendeeCount"]
        total = sum(types.values())
        for t, count in sorted(types.items(), key=lambda x: -x[1]):
            reg_type_data.append({
                "year": str(y),
                "type": t,
                "count": count,
                "pct": round(count / total * 100, 1) if total else 0,
            })

    # --- Repeat attendance (company level) ---
    company_years = {}
    for r in all_records:
        c = r["company"].lower().strip()
        if c:
            company_years.setdefault(c, set()).add(r["year"])

    repeat_cohorts = {1: 0, 2: 0, 3: 0}
    for c, yrs in company_years.items():
        n = len(yrs)
        if n in repeat_cohorts:
            repeat_cohorts[n] += 1
        elif n > 3:
            repeat_cohorts[3] += 1

    repeat_data = [
        {"label": "1 Year Only", "count": repeat_cohorts[1]},
        {"label": "2 Years", "count": repeat_cohorts[2]},
        {"label": "3 Years", "count": repeat_cohorts[3]},
    ]

    # New vs returning companies per year
    new_vs_returning = []
    seen = set()
    for y in years:
        current = companies_by_year[y]
        returning = current & seen
        new = current - seen
        new_vs_returning.append({
            "year": str(y),
            "new": len(new),
            "returning": len(returning),
            "newPct": round(len(new) / len(current) * 100, 1) if current else 0,
            "returningPct": round(len(returning) / len(current) * 100, 1) if current else 0,
        })
        seen |= current

    # --- AM Performance (2024-2025 only) ---
    am_data = {}
    for r in all_records:
        if r["year"] < 2024 or not r["accountManager"]:
            continue
        am = r["accountManager"]
        y = r["year"]
        key = (am, y)
        if key not in am_data:
            am_data[key] = {"am": am, "year": str(y), "state": r["accountManagerState"], "registrations": 0, "companies": set(), "venues": set()}
        am_data[key]["registrations"] += 1
        am_data[key]["companies"].add(r["company"].lower())
        am_data[key]["venues"].add(r["venue"])

    am_performance = []
    for key, d in sorted(am_data.items()):
        am_performance.append({
            "am": d["am"],
            "year": d["year"],
            "state": d["state"],
            "registrations": d["registrations"],
            "uniqueCompanies": len(d["companies"]),
            "venues": sorted(d["venues"]),
        })

    # AM summary across years
    am_summary = {}
    for entry in am_performance:
        am = entry["am"]
        if am not in am_summary:
            am_summary[am] = {"am": am, "state": entry["state"], "years": {}, "totalRegistrations": 0, "totalCompanies": set()}
        am_summary[am]["years"][entry["year"]] = entry["registrations"]
        am_summary[am]["totalRegistrations"] += entry["registrations"]

    # Collect unique companies per AM across all years
    for r in all_records:
        if r["year"] >= 2024 and r["accountManager"] and r["accountManager"] in am_summary:
            am_summary[r["accountManager"]]["totalCompanies"].add(r["company"].lower())

    am_summary_list = []
    for am, d in sorted(am_summary.items(), key=lambda x: -x[1]["totalRegistrations"]):
        am_summary_list.append({
            "am": d["am"],
            "state": d["state"],
            "totalRegistrations": d["totalRegistrations"],
            "uniqueCompanies": len(d["totalCompanies"]),
            "by2024": d["years"].get("2024", 0),
            "by2025": d["years"].get("2025", 0),
            "growth": round((d["years"].get("2025", 0) - d["years"].get("2024", 0)) / d["years"].get("2024", 1) * 100, 1) if d["years"].get("2024", 0) else None,
        })

    # --- Venue attendance by year (for city comparison) ---
    venue_by_year = {}
    for y in years:
        for r in by_year[y]:
            v = r["venue"] or "Unknown"
            venue_by_year.setdefault(v, {})[str(y)] = venue_by_year.get(v, {}).get(str(y), 0) + r["attendeeCount"]

    venue_comparison = []
    for v, year_counts in sorted(venue_by_year.items()):
        entry = {"venue": v}
        entry.update(year_counts)
        venue_comparison.append(entry)

    return {
        "summary": {
            "totalRecords": len(all_records),
            "years": [str(y) for y in years],
            "totalCompanies": len(company_years),
        },
        "yearOverYear": yoy,
        "venueData": venue_data,
        "venueComparison": venue_comparison,
        "registrationTypes": reg_type_data,
        "repeatAttendance": repeat_data,
        "newVsReturning": new_vs_returning,
        "amPerformance": am_performance,
        "amSummary": am_summary_list,
    }


def main():
    print("Processing 2023 data...")
    records_2023 = read_2023()
    print(f"  -> {len(records_2023)} records")

    print("Processing 2024 data...")
    records_2024 = read_2024_2025(2024)
    print(f"  -> {len(records_2024)} records")

    print("Processing 2025 data...")
    records_2025 = read_2024_2025(2025)
    print(f"  -> {len(records_2025)} records")

    all_records = records_2023 + records_2024 + records_2025
    print(f"\nTotal: {len(all_records)} records")

    analytics = compute_analytics(all_records)

    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump(analytics, f, indent=2, ensure_ascii=False)
    print(f"\nOutput written to {OUT}")


if __name__ == "__main__":
    main()
